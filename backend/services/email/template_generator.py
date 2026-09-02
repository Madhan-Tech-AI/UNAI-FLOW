import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_recipient_template() -> bytes:
    """
    Generates a production-ready Excel workbook (.xlsx) template for recipient bulk upload.
    Includes:
    1. 'Recipients' sheet with required 'email' and optional 'name' columns.
    2. 'Instructions' sheet with comprehensive guidelines and examples.
    """
    wb = openpyxl.Workbook()

    # Define color scheme matching UNAI Flow design
    navy_fill = PatternFill(start_color="09101D", end_color="09101D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True, color="09101D")
    title_font = Font(name="Calibri", size=14, bold=True, color="2563EB")
    regular_font = Font(name="Calibri", size=11, color="334155")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # ── Sheet 1: Recipients ──
    ws_recipients = wb.active
    ws_recipients.title = "Recipients"
    ws_recipients.views.sheetView[0].showGridLines = True

    # Headers
    headers = ["email", "name", "company"]
    ws_recipients.append(headers)

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws_recipients.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws_recipients.row_dimensions[1].height = 26

    # Sample rows (real examples)
    sample_data = [
        ["alex.morgan@example.com", "Alex Morgan", "Apex Digital"],
        ["sarah.connor@example.com", "Sarah Connor", "Cyberdyne Systems"],
        ["david.beckham@example.com", "David Beckham", "Inter Miami"],
        ["elena.rostova@example.com", "Elena", "Quantum Tech"],
    ]

    for row_data in sample_data:
        ws_recipients.append(row_data)

    for r_idx in range(2, len(sample_data) + 2):
        ws_recipients.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(headers) + 1):
            cell = ws_recipients.cell(row=r_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Set column widths
    ws_recipients.column_dimensions["A"].width = 34
    ws_recipients.column_dimensions["B"].width = 24
    ws_recipients.column_dimensions["C"].width = 24

    # ── Sheet 2: Instructions ──
    ws_instructions = wb.create_sheet(title="Instructions")
    ws_instructions.views.sheetView[0].showGridLines = True

    ws_instructions.column_dimensions["A"].width = 6
    ws_instructions.column_dimensions["B"].width = 30
    ws_instructions.column_dimensions["C"].width = 50

    instructions_data = [
        ("", "UNAI Flow — Bulk Email Template Guide", ""),
        ("", "", ""),
        ("", "Column Name", "Requirements & Usage"),
        ("", "email (Required)", "Must be a valid email address format (e.g. user@example.com). Cannot be empty."),
        ("", "name (Optional)", "Recipient's name used for server-side personalization via {{name}} variable."),
        ("", "company (Optional)", "Optional extra attribute. Unused columns are safely ignored."),
        ("", "", ""),
        ("", "Key Guidelines", ""),
        ("", "1. Header Row", "Keep the first row exactly as: email, name"),
        ("", "2. Personalization", "Use {{name}} in your subject or body. If name is missing, it falls back to 'there'."),
        ("", "3. Duplicate Prevention", "Duplicate email rows in the same spreadsheet will be automatically deduplicated."),
        ("", "4. Suppression List", "Unsubscribed or bounced emails in your account will be skipped automatically."),
        ("", "5. File Limits", "Supports up to 10,000 recipients per campaign. Maximum file size is 10 MB."),
    ]

    for r_idx, (c1, c2, c3) in enumerate(instructions_data, 1):
        ws_instructions.append([c1, c2, c3])
        ws_instructions.row_dimensions[r_idx].height = 22

    # Style Title
    cell_title = ws_instructions.cell(row=1, column=2)
    cell_title.font = title_font

    # Style Header Row (row 3)
    for c_idx in (2, 3):
        cell = ws_instructions.cell(row=3, column=c_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

    # Style data rows (rows 4-6)
    for r_idx in range(4, 7):
        cell_b = ws_instructions.cell(row=r_idx, column=2)
        cell_c = ws_instructions.cell(row=r_idx, column=3)
        cell_b.font = bold_font
        cell_c.font = regular_font
        cell_b.border = thin_border
        cell_c.border = thin_border

    # Style guidelines title (row 8)
    ws_instructions.cell(row=8, column=2).font = Font(name="Calibri", size=12, bold=True, color="2563EB")

    # Output bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
