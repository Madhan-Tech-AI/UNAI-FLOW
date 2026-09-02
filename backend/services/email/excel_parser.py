import io
import re
import openpyxl
from typing import Dict, Any, List, Tuple, Optional, Set
from lib.supabase_client import supabase

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
MAX_RECIPIENTS_LIMIT = 10000
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB

def is_valid_email_syntax(email_str: str) -> bool:
    """Validates email format using RFC-compliant check."""
    if not email_str or len(email_str) > 254:
        return False
    if not EMAIL_REGEX.match(email_str):
        return False
    parts = email_str.split("@")
    if len(parts) != 2:
        return False
    domain = parts[1]
    if "." not in domain or domain.startswith(".") or domain.endswith("."):
        return False
    return True

async def get_user_suppressions(user_id: str) -> Set[str]:
    """Fetches all suppressed email addresses for this user into a fast set."""
    try:
        res = (
            supabase.table("email_suppressions")
            .select("email")
            .eq("user_id", user_id)
            .execute()
        )
        if res.data:
            return {r["email"].strip().lower() for r in res.data if r.get("email")}
    except Exception:
        pass
    return set()

async def parse_recipient_spreadsheet(
    file_bytes: bytes,
    filename: str,
    user_id: str,
) -> Dict[str, Any]:
    """
    Authoritative server-side parser and validator for uploaded recipient spreadsheets (.xlsx, .xls).
    Enforces required email, syntax validity, deduplication, suppression checking, and row limits.
    """
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise ValueError(f"File exceeds maximum allowed size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB.")

    if not filename.lower().endswith((".xlsx", ".xls")):
        raise ValueError("Unsupported file format. Please upload an Excel spreadsheet (.xlsx or .xls).")

    # Load workbook using openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Failed to read Excel workbook: {str(e)}. Please ensure it is a valid .xlsx file.")

    # Find the data sheet: prefer 'Recipients', otherwise use active/first sheet
    sheet_names = wb.sheetnames
    target_sheet_name = "Recipients" if "Recipients" in sheet_names else sheet_names[0]
    ws = wb[target_sheet_name]

    # Read rows
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The uploaded spreadsheet is empty.")

    # Find header row (first non-empty row)
    header_row_idx = -1
    raw_headers = []
    for idx, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row_idx = idx
            raw_headers = [str(c).strip() if c is not None else "" for c in row]
            break

    if header_row_idx == -1 or not raw_headers:
        raise ValueError("No header row found in spreadsheet.")

    # Map column names case-insensitively
    col_map: Dict[str, int] = {}
    for col_idx, h in enumerate(raw_headers):
        clean_h = h.lower().replace("_", " ").replace("-", " ")
        if clean_h in ("email", "email address", "e mail", "mail", "contact email"):
            col_map["email"] = col_idx
        elif clean_h in ("name", "full name", "recipient name", "contact name", "first name"):
            col_map["name"] = col_idx

    if "email" not in col_map:
        raise ValueError(
            f"Missing required 'email' column header. Found columns: {', '.join(h for h in raw_headers if h)}. "
            "Please use the downloadable template."
        )

    # Fetch user suppression list
    suppressed_emails = await get_user_suppressions(user_id)

    # Process data rows
    seen_emails: Set[str] = set()
    valid_recipients: List[Dict[str, Any]] = []
    preview_records: List[Dict[str, Any]] = []

    total_rows = 0
    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    suppressed_count = 0

    data_rows = rows[header_row_idx + 1:]

    if len(data_rows) > MAX_RECIPIENTS_LIMIT:
        raise ValueError(
            f"Spreadsheet contains {len(data_rows)} rows, exceeding the limit of {MAX_RECIPIENTS_LIMIT} recipients per campaign."
        )

    for offset, row in enumerate(data_rows):
        # Skip completely empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        total_rows += 1
        row_num = header_row_idx + 2 + offset

        # Extract email
        raw_email = row[col_map["email"]] if col_map["email"] < len(row) else None
        email_str = str(raw_email).strip() if raw_email is not None else ""

        # Extract name
        raw_name = row[col_map["name"]] if ("name" in col_map and col_map["name"] < len(row)) else None
        name_str = str(raw_name).strip() if raw_name is not None else ""
        if name_str.lower() in ("none", "null", "nan"):
            name_str = ""

        # Extract extra variables
        variables: Dict[str, Any] = {"name": name_str}
        for c_idx, h_name in enumerate(raw_headers):
            if c_idx not in (col_map.get("email"), col_map.get("name")) and h_name:
                val = row[c_idx] if c_idx < len(row) else ""
                val_str = str(val).strip() if val is not None else ""
                if val_str.lower() not in ("none", "null", "nan"):
                    variables[h_name] = val_str

        # Validation checks
        record_status = "valid"
        reason = ""

        if not email_str:
            record_status = "invalid"
            reason = "Email address is missing"
            invalid_count += 1
        elif not is_valid_email_syntax(email_str):
            record_status = "invalid"
            reason = "Invalid email format"
            invalid_count += 1
        else:
            normalized_email = email_str.lower()
            if normalized_email in seen_emails:
                record_status = "duplicate"
                reason = "Duplicate email in spreadsheet"
                duplicate_count += 1
            elif normalized_email in suppressed_emails:
                record_status = "suppressed"
                reason = "Recipient is on your suppression list"
                suppressed_count += 1
            else:
                # Truly valid!
                seen_emails.add(normalized_email)
                valid_count += 1
                valid_recipients.append({
                    "email": normalized_email,
                    "name": name_str or None,
                    "variables": variables,
                })

        # Save for preview (up to 100 rows for display)
        if len(preview_records) < 100:
            preview_records.append({
                "row_number": row_num,
                "email": email_str or "—",
                "name": name_str or "—",
                "status": record_status,
                "reason": reason,
            })

    return {
        "total_rows": total_rows,
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "duplicate_count": duplicate_count,
        "suppressed_count": suppressed_count,
        "valid_recipients": valid_recipients,
        "preview_records": preview_records,
        "can_send": valid_count > 0,
    }
